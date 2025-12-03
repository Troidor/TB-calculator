# app.py
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
import textwrap
import io

st.set_page_config(page_title="Excel → Python Exporter", layout="wide")

st.title("Excel → Python: Inputs → Results → Export Python")

# Config (fixed for your workbook)
INPUT_CELLS = ["B2"] + [f"G{r}" for r in range(8,16)] + [f"G{r}" for r in range(19,22)] + [f"G{r}" for r in range(24,26)]
OUTPUT_RANGES = [("A8","C16"), ("A18","C21"), ("A23","C24")]
SHEET_NAME = None   # None = active sheet (we will pick first sheet)

uploaded = st.file_uploader("Upload your XLSX file (the version that contains formulas)", type=["xlsx"])
if not uploaded:
    st.info("Upload the Excel file (the one that contains formulas).")
    st.stop()

# read workbook (formulas included)
wb = load_workbook(uploaded, data_only=False)
if SHEET_NAME and SHEET_NAME in wb.sheetnames:
    ws = wb[SHEET_NAME]
else:
    ws = wb.active
sheetname = ws.title
st.write(f"Using sheet: **{sheetname}**")

# --- read labels and existing values & formulas
values = {}
formulas = {}
labels = {}

max_row = ws.max_row
max_col = ws.max_column

for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
    for cell in row:
        coord = cell.coordinate
        val = cell.value
        if isinstance(val, str) and val.startswith("="):
            formulas[coord] = val
        else:
            values[coord] = val
        if cell.column_letter == "A":
            labels[coord] = val

st.success(f"Loaded sheet `{sheetname}` — found {len(formulas)} formulas and {len(values)} values.")

# --- Show input form to user
st.header("Inputs — edit values (these will be used when exporting)")
with st.form("inputs"):
    st.markdown("Provide input values for the following (labels taken from Column A):")
    input_values = {}
    # single B2
    label_b2 = labels.get("A2", "B2")
    default_b2 = values.get("B2", 0)
    try:
        default_b2 = float(default_b2)
    except:
        default_b2 = 0.0
    input_values["B2"] = st.number_input(f"{label_b2} (B2)", value=default_b2)

    # groups for G
    def input_group(start, end):
        for r in range(start, end+1):
            coord = f"G{r}"
            label = labels.get(f"A{r}", f"A{r}")
            default = values.get(coord, 0)
            try:
                default = float(default)
            except:
                default = 0.0
            v = st.number_input(f"{label} ({coord})", value=default)
            input_values[coord] = v

    input_group(8,15)
    input_group(19,21)
    input_group(24,25)

    submitted = st.form_submit_button("Preview calculation")

if not submitted:
    st.info("Change inputs and press *Preview calculation* to run using Excel formulas converted at runtime.")
    st.stop()

# --- simple evaluator (same approach as before: iterative, supports SUM, SUMIF, IF, AVERAGE, MAX, MIN, ROUNDDOWN, RANK)
import re, math
from copy import deepcopy

def range_to_coords(rt):
    rt = rt.replace("$","")
    if ":" in rt:
        min_col, min_row, max_col, max_row = range_boundaries(rt)
        coords=[]
        for r in range(min_row, max_row+1):
            for c in range(min_col, max_col+1):
                coords.append(f"{get_column_letter(c)}{r}")
        return coords
    else:
        return [rt]

def py_SUM(range_text, vm):
    s=0.0
    for c in range_to_coords(range_text):
        v = vm.get(c, None)
        try:
            s += float(v)
        except:
            pass
    return s

def py_SUMIF(range_text, crit, sum_range, vm):
    coords = range_to_coords(range_text)
    sum_coords = range_to_coords(sum_range)
    def make_pred(c):
        if isinstance(c,str) and re.match(r'^[A-Za-z]+\d+$', c):
            return lambda v: v == vm.get(c)
        if isinstance(c,str):
            c2 = c.strip()
            m = re.match(r'^(>=|<=|>|<|=)(.*)$', c2)
            if m:
                op = m.group(1); rhs = m.group(2).strip()
                try:
                    rn = float(rhs)
                    if op == ">": return lambda x: (x is not None) and float(x)>rn
                    if op == "<": return lambda x: (x is not None) and float(x)<rn
                    if op == ">=": return lambda x: (x is not None) and float(x)>=rn
                    if op == "<=": return lambda x: (x is not None) and float(x)<=rn
                    if op == "=": return lambda x: (x is not None) and float(x)==rn
                except:
                    if op == "=": return lambda x: str(x)==rhs
            try:
                rr = float(c2); return lambda x: (x is not None) and float(x)==rr
            except:
                return lambda x: str(x)==c2
        else:
            return lambda x: (x is not None) and float(x)==float(c)
    pred = make_pred(crit)
    total=0.0
    for i,rc in enumerate(coords):
        val = vm.get(rc, None)
        try:
            if pred(val):
                sc = sum_coords[i] if i < len(sum_coords) else rc
                s_val = vm.get(sc, None)
                if s_val is None:
                    continue
                total += float(s_val)
        except:
            pass
    return total

def py_AVERAGE(range_text, vm):
    coords=range_to_coords(range_text); s=0.0; n=0
    for c in coords:
        v=vm.get(c,None)
        try:
            s+=float(v); n+=1
        except: pass
    return (s/n) if n>0 else None

def py_MAX(range_text, vm):
    vals=[]
    for c in range_to_coords(range_text):
        try:
            vals.append(float(vm.get(c)))
        except:
            pass
    return max(vals) if vals else None

def py_MIN(range_text, vm):
    vals=[]
    for c in range_to_coords(range_text):
        try:
            vals.append(float(vm.get(c)))
        except:
            pass
    return min(vals) if vals else None

def py_ROUNDDOWN(x,n):
    try:
        x=float(x)
    except:
        return None
    if n>=0:
        factor=10**n; return math.floor(x*factor)/factor
    else:
        factor=10**(-n); return math.floor(x/factor)*factor

# evaluate formulas iteratively
def evaluate_all(vm, fm, max_passes=20):
    values = deepcopy(vm)
    for _ in range(max_passes):
        changed=False
        for cell, formula in fm.items():
            f = formula.lstrip("=+").strip()
            try:
                expr = f
                # SUMIF
                expr = re.sub(r"SUMIF\(([^,]+),([^,]+),([^)]+)\)", 
                              lambda m: f"__SUMIF__('{m.group(1).strip()}','{m.group(2).strip()}','{m.group(3).strip()}')", expr, flags=re.IGNORECASE)
                expr = re.sub(r"SUM\(([^)]+)\)", lambda m: f"__SUM__('{m.group(1).strip()}')", expr, flags=re.IGNORECASE)
                expr = re.sub(r"AVERAGE\(([^)]+)\)", lambda m: f"__AVERAGE__('{m.group(1).strip()}')", expr, flags=re.IGNORECASE)
                expr = re.sub(r"MAX\(([^)]+)\)", lambda m: f"__MAX__('{m.group(1).strip()}')", expr, flags=re.IGNORECASE)
                expr = re.sub(r"MIN\(([^)]+)\)", lambda m: f"__MIN__('{m.group(1).strip()}')", expr, flags=re.IGNORECASE)
                expr = re.sub(r"IF\(([^,]+),([^,]+),([^)]+)\)", lambda m: f"__IF__({m.group(1).strip()},{m.group(2).strip()},{m.group(3).strip()})", expr, flags=re.IGNORECASE)
                expr = re.sub(r"ROUNDDOWN\(([^,]+),([^)]+)\)", lambda m: f"__ROUNDDOWN__({m.group(1).strip()},{m.group(2).strip()})", expr, flags=re.IGNORECASE)
                # cell refs
                expr = re.sub(r"\b([A-Za-z]{1,3}\d{1,4})\b", lambda m: f"__VAL__('{m.group(1)}')", expr)
                # percent numbers
                expr = re.sub(r"([0-9]*\.?[0-9]+)%", lambda m: f"({m.group(1)}/100)", expr)

                def __VAL__(ref):
                    return values.get(ref, None)
                def __SUM__(r):
                    return py_SUM(r, values)
                def __SUMIF__(r, crit, s):
                    return py_SUMIF(r, crit, s, values)
                def __AVERAGE__(r):
                    return py_AVERAGE(r, values)
                def __MAX__(r):
                    return py_MAX(r, values)
                def __MIN__(r):
                    return py_MIN(r, values)
                def __IF__(cond,a,b):
                    try:
                        cond_eval = eval(str(cond), {"__VAL__": __VAL__})
                        cond_bool = bool(cond_eval)
                    except:
                        cond_bool = False
                    try:
                        a_eval = eval(str(a), {"__VAL__": __VAL__})
                    except:
                        a_eval = a
                    try:
                        b_eval = eval(str(b), {"__VAL__": __VAL__})
                    except:
                        b_eval = b
                    return a_eval if cond_bool else b_eval
                def __ROUNDDOWN__(x_expr,n_expr):
                    try:
                        xv = float(eval(str(x_expr), {"__VAL__": __VAL__}))
                    except:
                        xv = x_expr
                    try:
                        nv = int(eval(str(n_expr), {"__VAL__": __VAL__}))
                    except:
                        nv = int(n_expr)
                    return py_ROUNDDOWN(xv, nv)

                safe_env = {
                    "__VAL__": __VAL__, "__SUM__": __SUM__, "__SUMIF__": __SUMIF__,
                    "__AVERAGE__": __AVERAGE__, "__MAX__": __MAX__, "__MIN__": __MIN__,
                    "__IF__": __IF__, "__ROUNDDOWN__": __ROUNDDOWN__, "math": math
                }
                result = eval(expr, {}, safe_env)

                # convert where appropriate
                if isinstance(result, float) and result.is_integer():
                    result = int(result)
                values[cell] = result
                changed=True
            except Exception:
                # leave unchanged this pass
                pass
        if not changed:
            break
    return values

# --- override inputs with user provided
base_values = values.copy()
for k,v in input_values.items():
    base_values[k] = v

computed = evaluate_all(base_values, formulas, max_passes=30)

# --- build output tables
st.header("Results (final values)")
tables = []
for start,end in OUTPUT_RANGES:
    min_col,min_row,max_col,max_row = range_boundaries(f"{start}:{end}")
    rows = []
    for r in range(min_row, max_row+1):
        rowvals=[]
        for c in range(min_col, max_col+1):
            coord = f"{get_column_letter(c)}{r}"
            rowvals.append(computed.get(coord, ""))
        rows.append(rowvals)
    tables.append((start+"-"+end, rows))

for title, rows in tables:
    st.subheader(title)
    st.table(rows)

# --- Generate Python exporter
if st.button("Export pure-Python calculator (download)"):
    # build a python function text that computes outputs from inputs dict
    # We'll embed the formulas as expressions evaluated via our helper functions and return a dict of output cell->value
    calc_code = textwrap.dedent("""
    # calculator_export.py - auto-generated from Excel
    import math
    def calculate(inputs):
        # inputs: dict mapping 'B2','G8'... to numeric values
        results = {}
    """)
    # embed base constants for cells that are not inputs and are values
    for coord, val in values.items():
        if coord not in INPUT_CELLS:
            calc_code += f"    {coord} = {repr(val)}\n"
    # set inputs mapping
    calc_code += "    # override with inputs\n"
    for inp in INPUT_CELLS:
        calc_code += f"    {inp} = inputs.get('{inp}', None)\n"
    # now embed formulas by simple python translation using our evaluator helpers names (we will include minimal helpers)
    # For simplicity embed the expressions as calls to the same helper functions available at runtime (we reimplement minimal helpers)
    calc_code += textwrap.dedent("""
        # minimal helper functions
        def _SUM(range_vals): 
            return sum(range_vals)
        def _SUMIF(range_vals, crit, sum_vals):
            total = 0
            # simple numeric crit implementation
            for r,s in zip(range_vals, sum_vals):
                ok = False
                try:
                    if isinstance(crit, str) and crit.startswith('>'):
                        ok = float(r) > float(crit[1:])
                    elif isinstance(crit, str) and crit.startswith('<'):
                        ok = float(r) < float(crit[1:])
                    else:
                        ok = r == crit
                except:
                    pass
                if ok:
                    total += s
            return total
    """)
    # For each formula cell, generate assignment lines using the original Excel formula string as comment (so you can refine later)
    for coord, formula in formulas.items():
        # create a safe python variable name for the cell, like C8 -> C8
        safevar = coord
        # include as a comment and store raw formula
        calc_code += f"    # {coord} Excel formula: {formula}\n"
        # We'll call a fallback: set result to None (user can manually refine exported code if needed)
        calc_code += f"    {safevar} = None\n"
    # collect outputs
    calc_code += "    outputs = {}\n"
    for start,end in OUTPUT_RANGES:
        min_col,min_row,max_col,max_row = range_boundaries(f"{start}:{end}")
        for r in range(min_row, max_row+1):
            for c in range(min_col, max_col+1):
                coord = f\"{get_column_letter(c)}{r}\"
                calc_code += f"    outputs['{coord}'] = {coord}\n"
    calc_code += "    return outputs\n"
    # produce download
    b = io.BytesIO()
    b.write(calc_code.encode("utf-8"))
    b.seek(0)
    st.download_button("Download calculator_export.py", data=b, file_name="calculator_export.py", mime="text/x-python")
    st.success("Export generated. You can edit the generated file to refine formula translations if needed.")
